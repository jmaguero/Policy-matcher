import json

from datetime import datetime
from pathlib import Path

import openpyxl

from llm import call_llm

OUTPUTS_DIR = Path(__file__).parent / "outputs"
REQUIRED_KEYS = ["rewritten_suggestions"]

# LLM05: output value constraints
MAX_SUGGESTION_ITEMS = 10
MAX_SUGGESTION_CHARS = 2_000


def _read_xlsx_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    return [
        dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)
    ]


def _build_user_prompt(row: dict) -> str:
    # LLM01: wrap the suggestions field (LLM-generated from Analyze button) in
    # explicit structural delimiters — prevents second-order prompt injection
    # when chaining LLM outputs as inputs.
    return (
        f"Control ID: {row['id']}\n"
        f"Title: {row['title']}\n"
        f"Control requirement: {row['control']}\n\n"
        f"<suggestions>\n{row['suggestions']}\n</suggestions>"
    )


def _resolve_input(filename: str) -> Path:
    """Resolve filename to an absolute path inside OUTPUTS_DIR.

    Uses Path.is_relative_to() to avoid the startswith() prefix-collision bug
    (e.g. /app/outputs2 falsely matching /app/outputs).
    """
    resolved = (OUTPUTS_DIR / filename).resolve()
    if not resolved.is_relative_to(OUTPUTS_DIR.resolve()):
        raise ValueError("Invalid input filename")
    if not resolved.exists():
        raise FileNotFoundError(f"File not found in outputs: {filename}")
    return resolved


def _validate_llm_output(result: dict) -> list[str]:
    """LLM05: enforce list type, item count, and per-item length."""
    rewritten = result.get("rewritten_suggestions", [])
    if isinstance(rewritten, str):
        rewritten = [rewritten]
    if not isinstance(rewritten, list):
        raise ValueError("rewritten_suggestions must be a list")
    # Cap item count and each item's length
    rewritten = rewritten[:MAX_SUGGESTION_ITEMS]
    return [str(item)[:MAX_SUGGESTION_CHARS] for item in rewritten]


def rewrite_suggestions(
    input_xlsx_filename: str,
    system_prompt: str,
    provider: str,
    model: str,
) -> dict:
    OUTPUTS_DIR.mkdir(exist_ok=True)

    input_path = _resolve_input(input_xlsx_filename)
    rows = _read_xlsx_rows(input_path)

    results = []
    for row in rows:
        row_id = str(row.get("id", ""))
        suggestions = (row.get("suggestions") or "").strip()

        if not suggestions:
            results.append(
                {
                    "id": row_id,
                    "rewritten_suggestions": [],
                    "selected_llm2": model,
                }
            )
            continue

        user_prompt = _build_user_prompt(row)
        llm_result = call_llm(
            provider=provider,
            model=model,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            required_keys=REQUIRED_KEYS,
        )
        rewritten = _validate_llm_output(llm_result)
        results.append(
            {
                "id": row_id,
                "rewritten_suggestions": rewritten,
                "selected_llm2": model,
            }
        )

    # Carry the client+standard part forward from the input filename.
    # Input stem format: <YYYYMMDD>_<HHMMSS>_<client>_<standard>
    stem = Path(input_xlsx_filename).stem
    parts = stem.split("_", 2)
    remainder = parts[2] if len(parts) > 2 else stem

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = OUTPUTS_DIR / f"{ts}_output_llm2.json"
    json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False))

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    result_by_id = {str(r["id"]): r for r in results}

    new_cols = ["rewritten_suggestions", "selected_llm2"]
    last_col = ws.max_column
    for i, col_name in enumerate(new_cols, start=1):
        ws.cell(row=1, column=last_col + i, value=col_name)

    for row_idx in range(2, ws.max_row + 1):
        row_id = str(ws.cell(row=row_idx, column=1).value)
        r = result_by_id.get(row_id)
        if r:
            for i, col_name in enumerate(new_cols, start=1):
                value = r[col_name]
                if isinstance(value, list):
                    value = "\n".join(value)
                ws.cell(row=row_idx, column=last_col + i, value=value)

    xlsx_out_path = OUTPUTS_DIR / f"{ts}_{remainder}.xlsx"
    wb.save(xlsx_out_path)

    return {
        "json_file": json_path.name,
        "xlsx_file": xlsx_out_path.name,
        "results": results,
    }
