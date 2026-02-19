import io
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber

from llm import call_llm

OUTPUTS_DIR = Path(__file__).parent / "outputs"
REQUIRED_KEYS = ["match", "if_yes_reason", "suggestions"]

# LLM10: resource limits
MAX_PDF_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_PDF_TEXT_CHARS = 50_000  # ~12k tokens — enough for any real policy doc

# LLM05: output value constraints
ALLOWED_MATCH_VALUES = {"yes", "no", "partial"}
MAX_FIELD_CHARS = 4_000


def _sanitize(s: str) -> str:
    """Keep only safe characters for use in a filename component."""
    return re.sub(r"[^\w\-]", "_", s).strip("_")


def _validate_pdf(pdf_bytes: bytes) -> None:
    """Reject oversized uploads and non-PDF magic bytes."""
    if len(pdf_bytes) > MAX_PDF_BYTES:
        raise ValueError(
            f"PDF exceeds maximum allowed size ({MAX_PDF_BYTES // 1024 // 1024} MB)"
        )
    if not pdf_bytes.startswith(b"%PDF"):
        raise ValueError("Uploaded file is not a valid PDF")


def _extract_pdf_text(pdf_bytes: bytes) -> str:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            pages = [page.extract_text() or "" for page in pdf.pages]
    except Exception:
        raise ValueError("Could not extract text from the uploaded PDF")
    text = "\n\n".join(pages)
    # LLM10: cap the text that goes into every prompt
    return text[:MAX_PDF_TEXT_CHARS]


def _read_xlsx_rows(xlsx_bytes: bytes) -> list[dict]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    except Exception:
        raise ValueError("Could not open the uploaded XLSX file")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    return [
        dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)
    ]


def _build_user_prompt(row: dict, pdf_text: str) -> str:
    # LLM01: wrap untrusted content in explicit structural delimiters so the
    # model treats it as data, not as instructions.
    return (
        f"Control ID: {row['id']}\n"
        f"Title: {row['title']}\n"
        f"Control requirement: {row['control']}\n\n"
        f"<policy_document>\n{pdf_text}\n</policy_document>"
    )


def _validate_llm_output(result: dict) -> dict:
    """LLM05: enforce expected types and value constraints on LLM output."""
    match_val = str(result.get("match", "")).strip().lower()
    if match_val not in ALLOWED_MATCH_VALUES:
        raise ValueError(f"LLM returned unexpected match value: {match_val!r}")

    return {
        "match": match_val,
        "if_yes_reason": str(result.get("if_yes_reason", ""))[:MAX_FIELD_CHARS],
        "suggestions": str(result.get("suggestions", ""))[:MAX_FIELD_CHARS],
    }


def analyze_policy(
    pdf_bytes: bytes,
    xlsx_bytes: bytes,
    xlsx_filename: str,
    client_name: str,
    system_prompt: str,
    provider: str,
    model: str,
) -> dict:
    OUTPUTS_DIR.mkdir(exist_ok=True)

    _validate_pdf(pdf_bytes)
    pdf_text = _extract_pdf_text(pdf_bytes)
    rows = _read_xlsx_rows(xlsx_bytes)

    results = []
    for row in rows:
        user_prompt = _build_user_prompt(row, pdf_text)
        llm_result = call_llm(
            provider=provider,
            model=model,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            required_keys=REQUIRED_KEYS,
        )
        validated = _validate_llm_output(llm_result)
        results.append(
            {
                "id": row["id"],
                "control": row["control"],
                "selected_llm1": model,
                **validated,
            }
        )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    standard = _sanitize(Path(xlsx_filename).stem)
    safe_client = _sanitize(client_name)

    json_path = OUTPUTS_DIR / f"{ts}_output_llm1.json"
    json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False))

    wb_orig = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws_orig = wb_orig.active
    result_by_id = {str(r["id"]): r for r in results}

    new_cols = ["match", "if_yes_reason", "suggestions", "selected_llm1"]
    last_col = ws_orig.max_column
    for i, col_name in enumerate(new_cols, start=1):
        ws_orig.cell(row=1, column=last_col + i, value=col_name)

    for row_idx in range(2, ws_orig.max_row + 1):
        row_id = str(ws_orig.cell(row=row_idx, column=1).value)
        r = result_by_id.get(row_id)
        if r:
            for i, col_name in enumerate(new_cols, start=1):
                ws_orig.cell(row=row_idx, column=last_col + i, value=r[col_name])

    xlsx_out_path = OUTPUTS_DIR / f"{ts}_{safe_client}_{standard}.xlsx"
    wb_orig.save(xlsx_out_path)

    return {
        "json_file": json_path.name,
        "xlsx_file": xlsx_out_path.name,
        "results": results,
    }
